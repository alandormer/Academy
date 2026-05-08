"""
Document text extraction.

Supports: PDF (PyMuPDF), DOCX (python-docx), XLSX (openpyxl),
XLS (xlrd), and plain text files.
Each parser returns a single cleaned string of extracted text.
"""
from __future__ import annotations

import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_text(filename: str, content: bytes) -> str:
    """
    Dispatch to the correct parser based on file extension.
    Returns extracted text as a single string.
    """
    suffix = Path(filename).suffix.lower()

    if suffix == ".pdf":
        return _extract_pdf(content)
    elif suffix == ".docx":
        return _extract_docx(content)
    elif suffix == ".xlsx":
        return _extract_xlsx(content)
    elif suffix == ".xls":
        return _extract_xls(content)
    elif suffix in {".txt", ".md", ".rst"}:
        return content.decode("utf-8", errors="replace")
    else:
        raise ValueError(f"Unsupported file type: {suffix!r}")


def _extract_pdf(content: bytes) -> str:
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:
        raise RuntimeError("pymupdf is required for PDF parsing. pip install pymupdf") from exc

    text_parts: list[str] = []
    with fitz.open(stream=content, filetype="pdf") as doc:
        for page in doc:
            text_parts.append(page.get_text())

    return "\n".join(text_parts)


def _extract_docx(content: bytes) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise RuntimeError("python-docx is required for DOCX parsing. pip install python-docx") from exc

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _extract_xlsx(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX parsing. pip install openpyxl") from exc

    parts: list[str] = []
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        parts.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                parts.append("\t".join(values))
        parts.append("")
    return "\n".join(parts).strip()


def _extract_xls(content: bytes) -> str:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required for XLS parsing. pip install xlrd") from exc

    parts: list[str] = []
    workbook = xlrd.open_workbook(file_contents=content)
    for idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(idx)
        parts.append(f"Sheet: {sheet.name}")
        for row_idx in range(sheet.nrows):
            row_values = []
            for col_idx in range(sheet.ncols):
                value = str(sheet.cell_value(row_idx, col_idx)).strip()
                if value:
                    row_values.append(value)
            if row_values:
                parts.append("\t".join(row_values))
        parts.append("")
    return "\n".join(parts).strip()
